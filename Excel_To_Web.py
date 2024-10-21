"""
Python 3.12.6

User Input:
    The PDF page number to read after executing the program.

Description:
    Extracts data from specified PDF pages and writes it to Excel.
    Each bibliographic entry consists of a title and an additional title,
    arranged in separate columns of a single row.

Developer:
    [Your Name]
"""

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

# Default book content checkbox for language selection
lang_path = "/html/body/div[2]/form/table/tbody/tr[4]/td/input[1]"

# List of data types for checkbox options on the web
data_types = {
    "Book": "/html/body/div/div/div/div/div[2]/table/tbody/tr[3]/td[2]/input[1]",
    "Reference Book": "/html/body/div/div/div/div/div[2]/table/tbody/tr[3]/td[2]/input[2]",
    "Journal Article": "/html/body/div/div/div/div/div[2]/table/tbody/tr[3]/td[2]/input[3]",
    "Book Review": "/html/body/div/div/div/div/div[2]/table/tbody/tr[3]/td[2]/input[4]",
    "Thesis": "/html/body/div/div/div/div/div[2]/table/tbody/tr[3]/td[2]/input[5]",
    "Conference Paper": "/html/body/div/div/div/div/div[2]/table/tbody/tr[3]/td[2]/input[6]",
    "Audio Material": "/html/body/div/div/div/div/div[2]/table/tbody/tr[3]/td[2]/input[7]",
    "Film and Video": "/html/body/div/div/div/div/div[2]/table/tbody/tr[3]/td[2]/input[8]",
    "Microform": "/html/body/div/div/div/div/div[2]/table/tbody/tr[3]/td[2]/input[9]",
    "Web Resource": "/html/body/div/div/div/div/div[2]/table/tbody/tr[3]/td[2]/input[10]",
    "Serial Publication": "/html/body/div/div/div/div/div[2]/table/tbody/tr[3]/td[2]/input[11]",
    "Proceedings": "/html/body/div/div/div/div/div[2]/table/tbody/tr[3]/td[2]/input[12]",
    "E-book": "/html/body/div/div/div/div/div[2]/table/tbody/tr[3]/td[2]/input[13]",
    "Research Paper": "/html/body/div/div/div/div/div[2]/table/tbody/tr[3]/td[2]/input[14]",
    "Other": "/html/body/div/div/div/div/div[2]/table/tbody/tr[3]/td[2]/input[15]",
}

# Open the browser
driver = webdriver.Edge()  # You can change this to your preferred browser

# Go to the library backend and log in(Replace with actual login URL)
driver.get('https://example-library-url.com/login')  

# Replace with actual username name and password name
username = driver.find_element(By.NAME, 'username')  
password = driver.find_element(By.NAME, 'password')  

# Replace with your username
# Replace with your password
username.send_keys('Your_Username')  
password.send_keys('Your_Password')  

# Click the login button
login_button = driver.find_element(By.NAME, 'submit')  
login_button.click()

time.sleep(2)

# Find the bibliographic management menu and click it
bib_management_menu = driver.find_element(By.XPATH, "/html/body/nav/div/div[2]/ul[1]/li[1]/a")  
bib_management_menu.click()

time.sleep(1)

# Find the "Create and Maintain Bibliography" option and click it
create_maintenance_option = driver.find_element(By.XPATH, "/html/body/nav/div/div[2]/ul[1]/li[1]/ul/li[1]/a")  
create_maintenance_option.click()

# Read data from the Excel file
def read_excel_data(excel_file, n):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active 

    data_list = [] 
    
    # Start reading data from the 2nd row in Excel
    for row in range(2, 2 + n):  
        entry = {
            "Title": sheet.cell(row=row, column=1).value,  
            "Additional Title": sheet.cell(row=row, column=2).value,  
            "Source Title": sheet.cell(row=row, column=3).value,  
            "Publisher": sheet.cell(row=row, column=4).value,  
            "Publisher URL": sheet.cell(row=row, column=5).value,  
            "Abstract": sheet.cell(row=row, column=6).value,  
            "Data Type": sheet.cell(row=row, column=7).value, 
            "Publication Year": sheet.cell(row=row, column=8).value,
            "Publication Month": sheet.cell(row=row, column=9).value,
            "Publication Day": sheet.cell(row=row, column=10).value,
            "Fulltext URL": sheet.cell(row=row, column=11).value,
        }
        data_list.append(entry)

    return data_list

# Use Selenium to automatically fill in the data
def fill_web_form(data, data_num): 
    wait = WebDriverWait(driver, 10)  

    # Find and click the create bibliography button
    create_button = driver.find_element(By.XPATH, "/html/body/div/div/div/div[1]/form/table/tbody/tr/td/div/input[1]")  
    create_button.click()

    # Input the title
    try:
        title_field = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "topic1")) 
        )
        
        driver.execute_script("arguments[0].value = arguments[1];", title_field, data[data_num]["Title"])
    except Exception as e:
        print(f"Error: {e}")

    # Uncheck all checkboxes
    for data_type, xpath in data_types.items():
        checkbox = driver.find_element(By.XPATH, xpath)
        if checkbox.is_selected():  
            checkbox.click()

    # Select the corresponding checkbox based on the "Data Type" from Excel for this entry
    checkbox_to_select = driver.find_element(By.XPATH, data_types[data[data_num]["Data Type"]])  
    checkbox_to_select.click()  

    # Click the check button to see if this entry already exists
    try:
        check_button = driver.find_element(By.ID, "checkBibliography")
        check_button.click()
        try:
            confirm_button = wait.until(EC.visibility_of_element_located((By.ID, "createBibliography")))
            confirm_button.click()  
        except:
            print("No option detected to continue creating; proceeding with other operations.")
    except Exception as e:
        print(f"Error occurred: {e}")

    time.sleep(1)  

    # Write the content from Excel into the corresponding column on the web
    topics_inputs = wait.until(EC.visibility_of_all_elements_located((By.NAME, "topics")))
    try:
        topics_inputs[1].send_keys(data[data_num]['Additional Title'])
        wait.until(EC.visibility_of_element_located((By.NAME, "srcTopic")))
        
        if data[data_num]['Source Title']:
            wait.until(EC.visibility_of_element_located((By.NAME, "srcTopic"))).send_keys(data[data_num]['Source Title'])
        else:
            print("Source Title not provided.")

        if data[data_num]['Publisher']:
            wait.until(EC.visibility_of_element_located((By.NAME, "publisher"))).send_keys(data[data_num]['Publisher'])
        else:
            print("Publisher not provided.")

        if data[data_num]['Publisher URL']:
            wait.until(EC.visibility_of_element_located((By.NAME, "publisherUrl"))).send_keys(data[data_num]['Publisher URL'])
        else:
            print("Publisher URL not provided.")
        
    except Exception as e:
        print(f"Error occurred: {e}")

    # Remove characters not in the Basic Multilingual Plane (BMP).
    def remove_non_bmp_characters(text):
        return ''.join(char for char in text if ord(char) <= 0xFFFF)
    
    # Filter the summary text to retain only BMP characters.
    summary_text = remove_non_bmp_characters(data[data_num]['Abstract'])
    # Set the filtered summary text in the "summary" input field using JavaScript.
    driver.execute_script("arguments[0].value = arguments[1];", driver.find_element(By.NAME, "summary"), summary_text)

    driver.find_element(By.NAME, "pressYear").send_keys(data[data_num]['Publication Year'])  
    driver.find_element(By.NAME, "pressMonth").send_keys(data[data_num]['Publication Month'])  
    driver.find_element(By.NAME, "pressDay").send_keys(data[data_num]['Publication Day']) 
    driver.find_element(By.NAME, "fullText").send_keys(data[data_num]['Fulltext URL'])  

    time.sleep(1)
    submit_button = driver.find_element(By.NAME, "submit")  # Submit the completed form
    submit_button.click()  

def main():
    n = int(input("Enter the number of bibliographic entries to process: "))  # Prompt the user for the number of entries
    data = read_excel_data("Bibliography_List_Example.xlsx", n)  # Replace with your Excel filename
    for i in range(n):  
        fill_web_form(data, i)

    driver.quit()  

if __name__ == "__main__":
    main()
